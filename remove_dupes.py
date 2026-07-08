from openpyxl import load_workbook
from openpyxl.styles import Alignment

def choose_file():
        filename = ''
        while not filename.endswith('.xlsx'):
            filename = input('Please Give The Excel Spreadsheet Containing The Grade Distribution: ')
            filename = filename.replace("'","")
            print(filename + ' selected')
            if not filename.endswith('.xlsx'):
                print('Please Give a valid .xlsx filename')
        return filename

spreadsheet_filename = choose_file()
wb = load_workbook(spreadsheet_filename)
ws = wb.active

def check_col_count():
    if ws.max_column != 21:
        print('this spreadsheet does not have the expected number of columns')
        print('script failed')
        exit()

def clean_dupes():
    dupe_count = 0
    unique_rows = []
    for row in ws.iter_rows(2, values_only = True):
        if row not in unique_rows:
            unique_rows.append(tuple(row))
        else:
            dupe_count += 1
    if dupe_count > 0:
        ws.delete_rows(2, ws.max_row)
        for unique_row in unique_rows:
            ws.append(unique_row)

        print(f'{dupe_count} duplicate rows found and deleted')

        wb.save(spreadsheet_filename)
    else: 
        print('no duplicate rows were found')

def remove_leading_zeros():
    for row in range(2, ws.max_row + 1):
        col_e = ws[f'E{row}'].value
        if isinstance(col_e, str) and col_e.isnumeric():
            ws[f'E{row}'] = int(col_e)
            ws[f'E{row}'].alignment = Alignment(horizontal='left')
    print('removed leading zeroes if there was any')        
    wb.save(spreadsheet_filename)

    
check_col_count()
clean_dupes()
remove_leading_zeros()