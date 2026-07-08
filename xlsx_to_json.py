import json, sys
from openpyxl import load_workbook
class qcdistroimport:
    def __init__(self):
        self.sheet= self.choose_file()

    def main(self):
        #D_plus only applys to years fall 2024 and up?
        column_names=['term','subject','course_number','course_desc','class_section','instructor',
                    'total_enrollment','A_plus','A','A_minus','B_plus','B','B_minus','C_plus','C','P',
                    'C_minus', 'D_plus','D','F',"Withdrawal","inc_ng","average_gpa"]
        
        output=[]
        wb = load_workbook(filename = self.sheet, read_only=True, data_only=True)
        sheet = self.choose_worksheet(wb)
        ws = wb[sheet]
        full_term = self.choose_term() + ' ' + str(self.choose_year())
        for row in ws[f'A2:{chr(ws.max_column + 64)}{ws.max_row}']:
            count=0
            column={}
            for value in row:  
                if value.value == None:
                    break
                if count == 6:
                    column[column_names[count]]= int(value.value)
                elif (count >=7 and count <= 20):
                        column[column_names[count]]= str(int(value.value) )
                elif (count == 2 or count ==4):
                    if type(value.value) is float:
                        column[column_names[count]]= str(int(value.value) )
                    else:
                        column[column_names[count]]= str(value.value)
                
                
                else:
                    if (count == 21 ):
                        column[column_names[count]]= int(value.value)
                    else:
                        column[column_names[count]]= str(value.value)
                
                count+=1

                # if count == 22:
                #     column[column_names[count]]= str(value.value)
                if count == 15:
                    column[column_names[count]]='0'
                    count+=1

                
            if len(column) == 0:
                continue
            column["term"] = full_term
            output.append(column)
        result_json = json.dumps(output, indent=4)
        self.write_to_file('qc {0}.json'.format(full_term.lower()),result_json)

    def write_to_file(self,file_name,json_object):
        file = open(file_name,"w")
        file.write(json_object)
        file.close()
        print('created file',file_name)
    
    def choose_file(self):
        filename = ''
        while not filename.endswith('.xlsx'):
            filename = input('Please Give The Excel Spreadsheet Containing The Grade Distribution: ')
            filename = filename.replace("'","")
            print(filename + ' selected')
            if not filename.endswith('.xlsx'):
                print('Please Give a valid .xlsx filename')
        return filename
    
    def choose_worksheet(self, wb):
        chosen_sheet = ''
        print('worksheets: ' + (', ').join(wb.sheetnames))
        while chosen_sheet not in wb.sheetnames:
            chosen_sheet = input('Please choose a worksheet to import: ')
            if chosen_sheet not in wb.sheetnames:
                print('Please choose a valid worksheet!')
        return chosen_sheet
    def choose_term(self):
        chosen_term = ''
        
        while chosen_term != 'Fall' and chosen_term != 'Spring' and chosen_term != 'Summer':
            chosen_term = input('Is this the Fall, Spring, or Summer Semester? ')
            print(chosen_term + ' Selected')
            if chosen_term != 'Fall' and chosen_term != 'Spring' and chosen_term != 'Summer':
                print('Please only type Fall, Spring, or Summer!')
        return chosen_term
    
    def choose_year(self):
        chosen_year = 0

        if chosen_year < 2012:
            chosen_year = int(input('What year is this term in? '))
            if chosen_year < 2012:
                print('please give a valid year!')
        return chosen_year
        
if __name__ == '__main__':
    sheet = qcdistroimport()
    sheet.main()

#everything is string except for total enrollment and inc no grade!!!!